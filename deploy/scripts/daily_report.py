# -*- coding: utf-8 -*-
"""daily_report.py — รายงานวัตถุดิบที่ใช้ + ยอดขายรายเมนู ต่อวัน → export Excel ส่งบัญชี

ใช้:
  cd deploy
  py scripts/daily_report.py                 # วันนี้ (เวลาไทย)
  py scripts/daily_report.py --date 2026-06-13
  py scripts/daily_report.py --from 2026-06-01 --to 2026-06-13

ที่มาตัวเลข: เมนูที่ขาย (sales) × สูตร (recipes.qty_stock) → วัตถุดิบที่ใช้ + ต้นทุน
⚠ qty_stock ที่ยังเป็น 0 → วัตถุดิบรายการนั้นจะไม่ถูกคิด (กรอกในหน้า 'สูตรขนม' ก่อน)
ไฟล์ออก: deploy/reports/daily_report_<ช่วง>.xlsx
"""
import sys, os, json, urllib.request, urllib.parse, datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
DEPLOY = os.path.normpath(os.path.join(HERE, ".."))
BKK = dt.timezone(dt.timedelta(hours=7))


def load_env():
    env = {}
    path = os.path.join(DEPLOY, ".env.local")
    if os.path.exists(path):
        for ln in open(path, encoding="utf-8"):
            if "=" in ln and not ln.strip().startswith("#"):
                k, v = ln.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


ENV = load_env()
SB_URL = ENV.get("NEXT_PUBLIC_SUPABASE_URL")
SB_KEY = ENV.get("SUPABASE_SERVICE_ROLE_KEY") or ENV.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
if not SB_URL or not SB_KEY:
    sys.exit("✗ ขาด NEXT_PUBLIC_SUPABASE_URL / KEY ใน .env.local")


def sb(path):
    req = urllib.request.Request(
        f"{SB_URL}/rest/v1/{path}",
        headers={"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def arg(name, default=None):
    flag = f"--{name}"
    if flag in sys.argv:
        i = sys.argv.index(flag)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return default


def bkk_date(iso):
    d = dt.datetime.fromisoformat(iso.replace("Z", "+00:00"))
    return d.astimezone(BKK).date().isoformat()


def main():
    today = dt.datetime.now(BKK).date().isoformat()
    if arg("date"):
        d_from = d_to = arg("date")
    else:
        d_from = arg("from", today)
        d_to = arg("to", today)

    # UTC bounds (เผื่อขอบเขตวันเวลาไทย)
    lo = (dt.date.fromisoformat(d_from)).isoformat() + "T00:00:00+07:00"
    hi = (dt.date.fromisoformat(d_to) + dt.timedelta(days=1)).isoformat() + "T00:00:00+07:00"

    # ดึงยอดขาย + รายการ + เมนู
    sel = ("id,sold_at,total,source,channel,"
           "sale_items(quantity,unit_price,line_total,menu_item_id,menu_items(name,category))")
    q = f"sales?select={urllib.parse.quote(sel)}&sold_at=gte.{urllib.parse.quote(lo)}&sold_at=lt.{urllib.parse.quote(hi)}&order=sold_at"
    sales = sb(q)

    # ดึงสูตร (menu_item_id → ingredient + qty_stock)
    rsel = "menu_item_id,ingredient_id,qty_stock,ingredients(name,unit,avg_cost)"
    recipes = sb(f"recipes?select={urllib.parse.quote(rsel)}")
    rec_by_menu = {}
    for r in recipes:
        rec_by_menu.setdefault(r["menu_item_id"], []).append(r)

    # รวมผล
    menu_rows = {}   # (date, menu) -> {qty, revenue, category}
    ing_rows = {}    # (date, ingredient_id) -> {name, unit, qty, avg_cost}
    day_summary = {}  # date -> {revenue, bills}

    for s in sales:
        d = bkk_date(s["sold_at"])
        ds = day_summary.setdefault(d, {"revenue": 0.0, "bills": 0})
        ds["revenue"] += float(s.get("total") or 0)
        ds["bills"] += 1
        for it in s.get("sale_items") or []:
            qty = float(it.get("quantity") or 0)
            name = (it.get("menu_items") or {}).get("name") or f"#{it['menu_item_id']}"
            cat = (it.get("menu_items") or {}).get("category") or ""
            mk = (d, name)
            mr = menu_rows.setdefault(mk, {"qty": 0.0, "revenue": 0.0, "category": cat})
            mr["qty"] += qty
            mr["revenue"] += float(it.get("line_total") or 0)
            # วัตถุดิบที่ใช้ (ผ่านสูตร)
            for rc in rec_by_menu.get(it["menu_item_id"], []):
                qs = float(rc.get("qty_stock") or 0)
                if qs <= 0:
                    continue
                ing = rc.get("ingredients") or {}
                key = (d, rc["ingredient_id"])
                ir = ing_rows.setdefault(key, {
                    "name": ing.get("name") or rc["ingredient_id"],
                    "unit": ing.get("unit") or "",
                    "avg_cost": float(ing.get("avg_cost") or 0),
                    "qty": 0.0,
                })
                ir["qty"] += qty * qs

    write_xlsx(d_from, d_to, day_summary, menu_rows, ing_rows)


def write_xlsx(d_from, d_to, day_summary, menu_rows, ing_rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    HEAD = Font(bold=True, color="FFFFFF")
    FILL = PatternFill("solid", fgColor="2F5597")
    BAHT = "#,##0.00"

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.font = HEAD
            cell.fill = FILL
            cell.alignment = Alignment(horizontal="center")

    # 1) สรุปรายวัน
    ws = wb.active
    ws.title = "สรุปรายวัน"
    ws.append(["วันที่", "จำนวนบิล", "ยอดขายรวม", "ต้นทุนวัตถุดิบรวม"])
    cost_by_day = {}
    for (d, _), v in ing_rows.items():
        cost_by_day[d] = cost_by_day.get(d, 0.0) + v["qty"] * v["avg_cost"]
    for d in sorted(day_summary):
        ws.append([d, day_summary[d]["bills"], round(day_summary[d]["revenue"], 2), round(cost_by_day.get(d, 0.0), 2)])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for c in row:
            c.number_format = BAHT
    style_header(ws, 4)
    ws.column_dimensions["A"].width = 14
    for col in "BCD":
        ws.column_dimensions[col].width = 18

    # 2) วัตถุดิบที่ใช้ (สำหรับบัญชี)
    ws2 = wb.create_sheet("วัตถุดิบที่ใช้")
    ws2.append(["วันที่", "รหัส", "วัตถุดิบ", "หน่วย", "จำนวนที่ใช้", "ต้นทุน/หน่วย", "มูลค่ารวม"])
    for (d, ing_id) in sorted(ing_rows):
        v = ing_rows[(d, ing_id)]
        ws2.append([d, ing_id, v["name"], v["unit"], round(v["qty"], 4),
                    round(v["avg_cost"], 2), round(v["qty"] * v["avg_cost"], 2)])
    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=7):
        for c in row:
            c.number_format = BAHT
    style_header(ws2, 7)
    for col, w in {"A": 12, "B": 11, "C": 30, "D": 8, "E": 12, "F": 12, "G": 14}.items():
        ws2.column_dimensions[col].width = w
    if ws2.max_row == 1:
        ws2.append(["— ยังไม่มีข้อมูล: กรอก qty_stock ในหน้า 'สูตรขนม' ก่อน —"])

    # 3) ยอดขายรายเมนู
    ws3 = wb.create_sheet("ยอดขายรายเมนู")
    ws3.append(["วันที่", "เมนู", "หมวด", "จำนวน", "ยอดขาย"])
    for (d, name) in sorted(menu_rows):
        v = menu_rows[(d, name)]
        ws3.append([d, name, v["category"], round(v["qty"], 2), round(v["revenue"], 2)])
    for row in ws3.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.number_format = BAHT
    style_header(ws3, 5)
    for col, w in {"A": 12, "B": 32, "C": 16, "D": 10, "E": 14}.items():
        ws3.column_dimensions[col].width = w

    out_dir = os.path.join(DEPLOY, "reports")
    os.makedirs(out_dir, exist_ok=True)
    tag = d_from if d_from == d_to else f"{d_from}_{d_to}"
    path = os.path.join(out_dir, f"daily_report_{tag}.xlsx")
    wb.save(path)

    n_ing = len(ing_rows)
    n_menu = len(menu_rows)
    print(f"✓ บันทึก {path}")
    print(f"  ช่วง {d_from} ถึง {d_to} · เมนู {n_menu} แถว · วัตถุดิบ {n_ing} แถว")
    if n_ing == 0:
        print("  ⚠ วัตถุดิบยังว่าง — qty_stock เป็น 0 (กรอกในหน้า 'สูตรขนม' แล้วรันใหม่)")


if __name__ == "__main__":
    main()
