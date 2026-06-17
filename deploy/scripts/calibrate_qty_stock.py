# -*- coding: utf-8 -*-
"""calibrate_qty_stock.py — คำนวณ qty_stock อัตโนมัติจากข้อมูลจริงทั้งเดือน

วิธี: qty_stock (ถุงต่อถ้วย) = Σถุงพรีมิกซ์ที่ใช้ทั้งเดือน (จาก ร้านขนมแม่.xlsx)
                               ÷ Σถ้วยที่ขายทั้งเดือน (จาก FoodStory)
แล้ว PATCH ค่าลง recipes (เฉพาะลิงก์เมนูเดี่ยว FS-* ที่ pre-build ไว้)

ใช้:
  cd deploy
  py scripts/calibrate_qty_stock.py --month 2026-06 --dry   # ดูผลก่อน
  py scripts/calibrate_qty_stock.py --month 2026-06         # เขียนจริง
"""
import sys, os, json, urllib.request, urllib.parse, re
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEPLOY = os.path.normpath(os.path.join(HERE, ".."))
XLSX = os.path.join(DEPLOY, "..", "ร้านขนมแม่.xlsx")
DRY = "--dry" in sys.argv

def arg(name, d=None):
    f = f"--{name}"
    return sys.argv[sys.argv.index(f) + 1] if f in sys.argv and sys.argv.index(f) + 1 < len(sys.argv) else d

MONTH = arg("month", "2026-06")
YEAR, MON = MONTH.split("-")
SHEET = {"01": "ม.ค", "02": "ก.พ.", "03": "มี.ค.", "04": "เม.ย", "05": "พ.ค", "06": "มิ.ย"}.get(MON, "มิ.ย") + " " + str(int(YEAR) + 543)[-2:]

env = {}
for ln in open(os.path.join(DEPLOY, ".env.local"), encoding="utf-8"):
    if "=" in ln and not ln.strip().startswith("#"):
        k, v = ln.split("=", 1); env[k.strip()] = v.strip().strip('"').strip("'")
SB_URL = env["NEXT_PUBLIC_SUPABASE_URL"]; SB_KEY = env.get("SUPABASE_SERVICE_ROLE_KEY") or env["NEXT_PUBLIC_SUPABASE_ANON_KEY"]
FB = "https://owner.foodstory.co"

# ── ถุงพรีมิกซ์ (ingredient_id) → ชื่อแถวใน Excel ──
EXCEL_POUCH = {
    "ING-0001": "ข้าวเหนียวเปียกลำไย", "ING-0002": "เต้าส่วนมะพร้าวอ่อน",
    "ING-0003": "สาคูอัญชันมะพร้าวอ่อน", "ING-0004": "เหนียวเปียกดำ",
    "ING-0005": "ครองแครงมะพร้าวอ่อน", "ING-0006": "กล้วยขวชชี",
    "ING-0072": "ข้าวเหนียวถั่วดำ", "ING-0023": "บัวลอยอัญชันลูกเล็ก",
    "ING-0022": "บัวลอยมีไส้",
}

def clean_name(name, code):
    name = (name or "").strip()
    if code and name.startswith(str(code) + " "):
        return name[len(str(code)) + 1:].strip()
    return name

def sb(path, method="GET", body=None, prefer=None):
    h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type": "application/json"}
    if prefer: h["Prefer"] = prefer
    req = urllib.request.Request(f"{SB_URL}/rest/v1/{path}", method=method,
                                 data=json.dumps(body).encode() if body else None, headers=h)
    with urllib.request.urlopen(req, timeout=30) as r:
        t = r.read().decode()
        return json.loads(t) if t else None

def fb_post(path, body):
    h = {"Cookie": env["FOODSTORY_COOKIE"], "X-CSRF-Token": env["FOODSTORY_CSRF"],
         "X-Requested-With": "XMLHttpRequest", "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
         "Origin": FB, "Referer": f"{FB}/th/salebyproductdaily",
         "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"}
    req = urllib.request.Request(FB + path, method="POST",
                                 data=urllib.parse.urlencode(body).encode(), headers=h)
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read().decode()

def fetch_foodstory_cups():
    fb_post("/api/setTimeLenght", {"year": YEAR})
    fb_post("/api/setMonthLenght", {"month": str(int(MON))})
    cols = {"draw": "1", "start": "0", "length": "9000", "order[0][column]": "0", "order[0][dir]": "asc"}
    names = ["show_dt", "product_code", "product_name", "sales_volumn"]
    for i, c in enumerate(names):
        cols[f"columns[{i}][data]"] = c; cols[f"columns[{i}][name]"] = c
        cols[f"columns[{i}][searchable]"] = "true"; cols[f"columns[{i}][orderable]"] = "true"
    data = json.loads(fb_post("/salebyproductdaily/getdata", cols))
    cups = {}
    for r in data.get("data", []):
        nm = clean_name(r.get("product_name"), r.get("product_code"))
        cups[nm] = cups.get(nm, 0) + float(r.get("sales_volumn") or 0)
    return cups, len(data.get("data", []))

def excel_pouch_usage():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    # คอลัมน์ "จำนวน" รายวัน = col 6,9,12,... (ทุก 3 เริ่ม F) ที่มี date header ใน R1
    day_cols = [c for c in range(6, ws.max_column + 1, 3) if ws.cell(1, c).value is not None]
    def num(v):
        try: return float(v)
        except (TypeError, ValueError): return 0.0
    usage = {}  # row label → Σ
    for r in range(3, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if not label: continue
        s = sum(num(ws.cell(r, c).value) for c in day_cols)
        usage[str(label).strip()] = s
    return usage

def main():
    print(f"calibrate qty_stock · เดือน {MONTH} (ชีต Excel: {SHEET}){'  [DRY]' if DRY else ''}")
    cups, nrows = fetch_foodstory_cups()
    print(f"  FoodStory: {nrows} แถว · {len(cups)} เมนู")
    usage = excel_pouch_usage()

    # ลิงก์ recipe ของเมนู FS-* (menu_item_id → ingredient_id + ชื่อเมนู)
    links = sb("recipes?select=id,menu_item_id,ingredient_id,menu_items!inner(menu_id,name)&menu_items.menu_id=like.FS-*")

    # รวมข้อมูลต่อ ingredient_id
    per_ing = {}  # ing_id → {cups, link_ids:[], menu_names:set}
    for l in links:
        ing = l["ingredient_id"]; nm = (l["menu_items"] or {}).get("name", "")
        d = per_ing.setdefault(ing, {"cups": 0.0, "ids": [], "names": set()})
        d["ids"].append(l["id"])
        if nm not in d["names"]:
            d["names"].add(nm); d["cups"] += cups.get(nm, 0)

    def pouch_sum(ing_id):
        key = EXCEL_POUCH.get(ing_id)
        if not key: return None
        for label, s in usage.items():
            if key in label:
                return s
        return None

    print(f"\n{'ingredient':<11}{'ถุงใช้':>8}{'ถ้วยขาย':>9}{'qty_stock':>11}  เมนู")
    updates = []
    for ing, d in per_ing.items():
        pu = pouch_sum(ing)
        if pu is None or d["cups"] <= 0 or pu <= 0:
            print(f"{ing:<11}{(pu if pu is not None else '-'):>8}{d['cups']:>9.0f}{'ข้าม':>11}  ({', '.join(d['names'])})")
            continue
        qty = round(pu / d["cups"], 4)
        print(f"{ing:<11}{pu:>8.0f}{d['cups']:>9.0f}{qty:>11.4f}  ({', '.join(d['names'])})")
        for rid in d["ids"]:
            updates.append((rid, qty))

    if DRY:
        print(f"\n[DRY] จะอัปเดต {len(updates)} ลิงก์ — ไม่เขียน")
        return
    for rid, qty in updates:
        sb(f"recipes?id=eq.{rid}", method="PATCH", body={"qty_stock": qty}, prefer="return=minimal")
    print(f"\n✓ อัปเดต qty_stock {len(updates)} ลิงก์แล้ว → ตัดสต็อก/รายงานวัตถุดิบใช้ได้ทันที")

if __name__ == "__main__":
    main()
